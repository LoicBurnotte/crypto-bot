import asyncio
import csv
import io
import logging
import os
from contextlib import asynccontextmanager
from typing import Any, Literal, Optional

from dotenv import load_dotenv
from fastapi import FastAPI, HTTPException, Query, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, Response, StreamingResponse
from pydantic import BaseModel, Field

from config import cfg, PROFILES
from bot import TradingBot

load_dotenv()

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
logger = logging.getLogger(__name__)

bot = TradingBot()
bot_task: asyncio.Task | None = None


@asynccontextmanager
async def lifespan(app: FastAPI):
    global bot_task
    bot_task = asyncio.create_task(bot.run_loop(interval=10))
    logger.info("Bot loop started")
    yield
    bot.stop()
    if bot_task:
        bot_task.cancel()
        try:
            await bot_task
        except asyncio.CancelledError:
            pass


app = FastAPI(title="Crypto Trading Bot", version="3.0.0", lifespan=lifespan)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["GET", "POST", "PUT", "OPTIONS"],
    allow_headers=["*"],
)


@app.middleware("http")
async def force_cors(request: Request, call_next):
    if request.method == "OPTIONS":
        res = Response(status_code=204)
        res.headers["Access-Control-Allow-Origin"]  = "*"
        res.headers["Access-Control-Allow-Methods"] = "GET, POST, PUT, OPTIONS"
        res.headers["Access-Control-Allow-Headers"] = "*"
        return res
    response = await call_next(request)
    response.headers["Access-Control-Allow-Origin"] = "*"
    return response


# ── Helpers ───────────────────────────────────────────────────────────────────

def live(data) -> JSONResponse:
    res = JSONResponse(data)
    res.headers["Cache-Control"] = "no-store, no-cache, must-revalidate"
    return res


# ── Models ────────────────────────────────────────────────────────────────────

class TradeRequest(BaseModel):
    symbol:        str                    = Field(..., example="BTC/EUR")
    side:          Literal["buy", "sell"]
    amount_eur:    Optional[float]        = Field(None, gt=0)
    amount_crypto: Optional[float]        = Field(None, gt=0)


class WithdrawRequest(BaseModel):
    currency: str   = "EUR"
    amount:   float = Field(..., gt=0)
    key:      str   = Field(..., description="Kraken withdrawal key name")


class ConfigUpdate(BaseModel):
    symbols:            Optional[list[str]]  = None
    trade_amount_eur:   Optional[float]      = None
    trailing_stop_pct:  Optional[float]      = None
    take_profit_pct:    Optional[float]      = None
    rsi_period:         Optional[int]        = None
    rsi_oversold:       Optional[float]      = None
    rsi_overbought:     Optional[float]      = None
    rsi_timeframe:      Optional[str]        = None
    ema_fast:           Optional[int]        = None
    ema_slow:           Optional[int]        = None
    max_daily_loss_eur: Optional[float]      = None
    webhook_url:        Optional[str]        = None
    kraken_api_key:     Optional[str]        = None
    kraken_api_secret:  Optional[str]        = None


# ── Routes ────────────────────────────────────────────────────────────────────

@app.get("/")
def root():
    return {"status": "ok"}


@app.get("/health")
def health():
    return {"healthy": True}


# ── Config endpoints ──────────────────────────────────────────────────────────

@app.get("/config")
def get_config():
    """Returns current config. Credential values are masked (••••)."""
    return cfg.get(include_secrets=False)


@app.put("/config")
def update_config(body: ConfigUpdate):
    """
    Partially update config. Only non-None fields are applied.
    Changes take effect on the next bot tick (no restart needed).
    New symbols are seeded automatically.
    """
    data = {k: v for k, v in body.model_dump().items() if v is not None}
    if not data:
        raise HTTPException(status_code=400, detail="No fields to update")

    # Validate timeframe
    if "rsi_timeframe" in data and data["rsi_timeframe"] not in ("1h", "4h", "1d"):
        raise HTTPException(status_code=400, detail="rsi_timeframe must be 1h, 4h or 1d")

    cfg.update(data)

    # Reload exchange if credentials changed
    if "kraken_api_key" in data or "kraken_api_secret" in data:
        bot.reload_exchange()

    # Sync bot states for any new symbols
    bot.sync_symbols()

    logger.info("Config updated: %s", list(data.keys()))
    return cfg.get(include_secrets=False)


@app.post("/config/reset/{profile}")
def reset_config(profile: str):
    """Reset to a named profile: conservative, moderate, aggressive, custom."""
    try:
        result = cfg.reset(profile)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    bot.sync_symbols()
    logger.info("Config reset to profile: %s", profile)
    return result


@app.get("/config/profiles")
def get_profiles():
    """Returns all built-in profiles with their default values."""
    return PROFILES


# ── Status / portfolio ────────────────────────────────────────────────────────

@app.get("/status")
def get_status():
    return live({"assets": bot.get_status()})


@app.get("/portfolio")
async def get_portfolio():
    try:
        raw    = await asyncio.to_thread(bot.exchange.fetch_balance)
        _skip  = {"info", "free", "used", "total", "timestamp", "datetime"}

        def eur_price(currency: str) -> Optional[float]:
            if currency in ("EUR", "ZEUR"):
                return 1.0
            sym   = f"{currency}/EUR"
            state = bot.states.get(sym)
            if state and state.last_price:
                return state.last_price
            try:
                return float(bot.exchange.fetch_ticker(sym)["last"])
            except Exception:
                return None

        balances: dict = {}
        total_eur = 0.0

        for currency, info in raw.items():
            if not isinstance(info, dict) or currency in _skip:
                continue
            total = info.get("total") or 0
            if total <= 0:
                continue

            price     = eur_price(currency)
            eur_value = round(total * price, 2) if price is not None else None

            sym   = f"{currency}/EUR"
            state = bot.states.get(sym)
            unreal_pnl_eur = None
            if state and state.entry_price and state.last_price and total > 0:
                unreal_pnl_eur = round((state.last_price - state.entry_price) * info.get("free", 0), 2)

            balances[currency] = {
                "free":               round(info["free"]  or 0, 8),
                "used":               round(info["used"]  or 0, 8),
                "total":              round(total, 8),
                "eur_price":          round(price, 4) if price else None,
                "eur_value":          eur_value,
                "unrealised_pnl_eur": unreal_pnl_eur,
            }
            if eur_value is not None:
                total_eur += eur_value

        all_time_pnl = bot.trade_log.summary()["all_time_pnl"]

        return live({
            "balances":     balances,
            "total_eur":    round(total_eur, 2),
            "daily_pnl":    round(bot._daily_pnl, 2),
            "all_time_pnl": all_time_pnl,
            "paused_loss":  bot._paused_loss,
        })
    except Exception as e:
        logger.error("Portfolio fetch failed: %s", e)
        raise HTTPException(status_code=503, detail=str(e))


_EXPORT_COLUMNS = [
    ("timestamp",    "Timestamp"),
    ("symbol",       "Symbol"),
    ("side",         "Side"),
    ("price",        "Price (€)"),
    ("amount_eur",   "Amount (€)"),
    ("amount_crypto","Amount (crypto)"),
    ("pnl_eur",      "P&L (€)"),
    ("reason",       "Reason"),
    ("dry_run",      "Dry run"),
    ("order_id",     "Order ID"),
]


@app.get("/trades/years")
def trade_years():
    return {"years": bot.trade_log.available_years()}


@app.get("/trades/export")
def export_trades(
    year:   Optional[int] = Query(None, description="Filter to calendar year, e.g. 2025"),
    format: str           = Query("csv", pattern="^(csv|xlsx)$"),
):
    trades   = bot.trade_log.for_export(year)
    filename = f"trades{'_' + str(year) if year else '_all'}.{format}"

    if format == "csv":
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow([col for _, col in _EXPORT_COLUMNS])
        for t in trades:
            writer.writerow([t.get(key, "") for key, _ in _EXPORT_COLUMNS])
        buf.seek(0)
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{filename}"',
                     "Access-Control-Allow-Origin": "*",
                     "Access-Control-Expose-Headers": "Content-Disposition"},
        )

    # xlsx
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Trades {year or 'All'}"

    # Header row
    header_fill = PatternFill("solid", fgColor="1E2235")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, (_, col_label) in enumerate(_EXPORT_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, t in enumerate(trades, 2):
        for col_idx, (key, _) in enumerate(_EXPORT_COLUMNS, 1):
            val = t.get(key, "")
            if key == "dry_run":
                val = "Yes" if val else "No"
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    # Freeze header
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"',
                 "Access-Control-Allow-Origin": "*",
                 "Access-Control-Expose-Headers": "Content-Disposition"},
    )


@app.get("/trades")
def get_trades(
    limit:  int = Query(100, le=2000),
    offset: int = Query(0, ge=0),
):
    all_trades = bot.trade_log.all()   # newest first
    page = all_trades[offset: offset + limit]
    return live({
        "trades":  page,
        "total":   bot.trade_log.count(),
        "offset":  offset,
        "limit":   limit,
        **bot.trade_log.summary(),
    })


_TIMEFRAME_LIMITS = {
    "1h": ("1h", 24),
    "4h": ("4h", 42),
    "1d": ("1d", 30),
    "1w": ("1w", 52),
}


@app.get("/ohlcv")
async def get_ohlcv(
    symbol:    str = Query(..., example="BTC/EUR"),
    timeframe: str = Query("1h", pattern="^(1h|4h|1d|1w)$"),
    limit:     int = Query(0),
):
    if symbol not in cfg.symbols:
        raise HTTPException(status_code=400, detail=f"Unknown symbol '{symbol}'")
    tf, default_limit = _TIMEFRAME_LIMITS[timeframe]
    candles = await asyncio.to_thread(bot.fetch_ohlcv, symbol, tf, limit or default_limit)
    return live({"symbol": symbol, "timeframe": timeframe, "data": candles})


# ── Bot control ───────────────────────────────────────────────────────────────

@app.get("/bot/status")
def bot_status():
    return {
        "running":     bot._running,
        "dry_run":     cfg.dry_run,
        "daily_pnl":   round(bot._daily_pnl, 2),
        "paused_loss": bot._paused_loss,
    }


@app.post("/bot/start")
async def start_bot():
    global bot_task
    if not bot._running:
        bot_task = asyncio.create_task(bot.run_loop(interval=10))
    return {"running": True}


@app.post("/bot/stop")
async def stop_bot():
    global bot_task
    bot.stop()
    if bot_task and not bot_task.done():
        bot_task.cancel()
        try:
            await bot_task
        except asyncio.CancelledError:
            pass
    return {"running": False}


@app.post("/bot/symbol/{symbol:path}")
def set_symbol_state(symbol: str, disabled: bool):
    if symbol not in cfg.symbols:
        raise HTTPException(status_code=400, detail=f"Unknown symbol '{symbol}'")
    bot.states[symbol].disabled = disabled
    return {"symbol": symbol, "disabled": disabled}


@app.post("/bot/hold/{symbol:path}")
def set_hold_until_overbought(symbol: str, enabled: bool = True):
    if symbol not in cfg.symbols:
        raise HTTPException(status_code=400, detail=f"Unknown symbol '{symbol}'")
    bot.states[symbol].hold_until_overbought = enabled
    return {"symbol": symbol, "hold_until_overbought": enabled, "rsi_target": cfg.rsi_overbought}


@app.post("/bot/liquidate")
def liquidate_all():
    results = [{"symbol": s, **bot.execute_sell(s, reason="liquidate")} for s in cfg.symbols]
    return {"results": results}


@app.post("/trade")
def manual_trade(req: TradeRequest):
    if req.symbol not in cfg.symbols:
        raise HTTPException(status_code=400, detail=f"Unknown symbol '{req.symbol}'")
    result = (bot.execute_buy(req.symbol, req.amount_eur, reason="manual")
              if req.side == "buy"
              else bot.execute_sell(req.symbol, req.amount_crypto, reason="manual"))
    if not result.get("ok"):
        raise HTTPException(status_code=400, detail=result.get("error"))
    return result


@app.post("/withdraw")
def withdraw(req: WithdrawRequest):
    try:
        result = bot.exchange.withdraw(req.currency, req.amount, req.key)
        logger.info("Withdrawal: %.2f %s -> %s", req.amount, req.currency, req.key)
        return {"ok": True, "result": result}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
